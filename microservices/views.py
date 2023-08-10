import pandas as pd
import xlrd 
import openpyxl
from django.shortcuts import render
from .forms import ExcelUploadForm
from .models import User, Product

def upload_excel(request):
    if request.method == 'POST':
        print('POST request received')
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print('Form is valid')
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)

            # Process the parsed data and save it to the database
            for _, row in df.iterrows():
                username = row['username']
                password = row['password']
                privileges = row['privileges']
                email = row['email']
                refreshtoken = row['refreshtoken']

                print(f'Processing user: {username}')

                # Create a new User object and save it to the database
                user = User(
                    username=username,
                    password=password,
                    privileges=privileges,
                    email=email,
                    refreshtoken=refreshtoken
                )
                user.save()
                print(f'Saved user: {username}')

            return render(request, 'microservices/upload_success.html')

    else:
        form = ExcelUploadForm()

    print('Invalid form or GET request')
    return render(request, 'microservices/upload.html', {'form': form})

def upload_excel_products(request):
    if request.method == 'POST':
        print('POST request received')
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print('Form is valid')
            excel_file = request.FILES['excel_file']

            # Load the Excel file
            workbook = openpyxl.load_workbook(excel_file)

            # Select the first sheet in the workbook
            sheet = workbook.worksheets[0]

            # Assign hardcoded values for proveedor and marca
            proveedor = 'Truper'
            marca = 'Truper'

            # Find the header's row dynamically
            header_row = None
            required_fields = {'CODIGO', 'EAN', 'CATEGORIA', 'DESCRIPCION', 'PVP final'}
            for row in sheet.iter_rows():
                values = [cell.value for cell in row]
                print(f"Row values: {values}")

                if set(values) >= required_fields:
                    header_row = row
                    print("Header row found")
                    break

            if header_row is None:
                print('Header row not found in the Excel file')
                return render(request, 'microservices/upload.html', {'form': form})

            headers = [cell.value for cell in header_row]
            print(f"Headers: {headers}")

            # Process the parsed data and save it to the database
            column = 1  # Counter to track the position of rows in the Excel file
            
            for row in sheet.iter_rows(min_row=header_row[0].row + 1):  # Start from the row after the header
                values = [cell.value for cell in row]

                # Check if the row has only a single cell and skip it
                if len(values) == 1:
                    print(f'Skipping single-cell row at position: {column}')
                    column += 1
                    continue

                # Find the column indices based on column headers
                try:
                    sku_col = headers.index('CODIGO') + 1
                    ean_col = headers.index('EAN') + 1
                    categoria_col = headers.index('CATEGORIA') + 1
                    descripcion_col = headers.index('DESCRIPCION') + 1
                    pvp_col = headers.index('PVP final') + 1
                except ValueError:
                    print('Column headers not found in the Excel file')
                    return render(request, 'microservices/upload.html', {'form': form})

                sku = values[sku_col - 1] if sku_col else None
                ean = values[ean_col - 1] if ean_col else None
                categoria = values[categoria_col - 1] if categoria_col else None
                descripcion = values[descripcion_col - 1] if descripcion_col else None
                pvp = values[pvp_col - 1] if pvp_col else None

                print(f'Processing product: {sku}')

                # Check if any required field is empty and skip the row
                if any(
                    field is None or (isinstance(field, str) and field.strip() == '')
                    for field in (sku, ean, categoria, descripcion, pvp)
                ):
                    print(f'Skipping row with empty fields at position: {column}')
                    column += 1
                    continue

                # Create a new Product object and save it to the database
                product = Product(
                    sku=sku,
                    ean=ean,
                    proveedor=proveedor,
                    categoria=categoria,
                    marca=marca,
                    descripcion=descripcion,
                    pvp=pvp
                )
                product.save()
                print(f'Saved product: {sku}')

                column += 1

            return render(request, 'microservices/upload_success.html')

    else:
        form = ExcelUploadForm()

    print('Invalid form or GET request')
    return render(request, 'microservices/upload.html', {'form': form})

def upload_excel_productsTEST(request):
    if request.method == 'POST':
        print('POST request received')
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print('Form is valid')
            excel_file = request.FILES['excel_file']

            # Load the Excel file
            workbook = openpyxl.load_workbook(excel_file)

            # Select the first sheet in the workbook
            sheet = workbook.worksheets[0]

            # Assign hardcoded values for proveedor and marca
            proveedor = 'Truper'
            marca = 'Truper'

            # Find the header's row dynamically
            header_row = None
            required_fields = {'CODIGO', 'EAN', 'CATEGORIA', 'DESCRIPCION', 'PVP final'}
            for row in sheet.iter_rows():
                values = [cell.value for cell in row]
                print(f"Row values: {values}")

                if set(values) >= required_fields:
                    header_row = row
                    print("Header row found")
                    break

            if header_row is None:
                print('Header row not found in the Excel file')
                return render(request, 'microservices/upload.html', {'form': form})

            headers = [cell.value for cell in header_row]
            print(f"Headers: {headers}")

            # Process the parsed data and save it to the database
            column = 1  # Counter to track the position of rows in the Excel file
            
            for row in sheet.iter_rows(min_row=header_row[0].row + 1):  # Start from the row after the header
                values = [cell.value for cell in row]

                # Check if the row has only a single cell and skip it
                if len(values) == 1:
                    print(f'Skipping single-cell row at position: {column}')
                    column += 1
                    continue

                # Find the column indices based on column headers
                try:
                    sku_col = headers.index('CODIGO') + 1
                    ean_col = headers.index('EAN') + 1
                    categoria_col = headers.index('CATEGORIA') + 1
                    descripcion_col = headers.index('DESCRIPCION') + 1
                    pvp_col = headers.index('PVP final') + 1
                except ValueError:
                    print('Column headers not found in the Excel file')
                    return render(request, 'microservices/upload.html', {'form': form})

                sku = values[sku_col - 1] if sku_col else None
                ean = values[ean_col - 1] if ean_col else None
                categoria = values[categoria_col - 1] if categoria_col else None
                descripcion = values[descripcion_col - 1] if descripcion_col else None
                pvp = values[pvp_col - 1] if pvp_col else None

                print(f'Processing product: {sku}')

                # Check if any required field is empty and skip the row
                if any(
                    field is None or (isinstance(field, str) and field.strip() == '')
                    for field in (sku, ean, categoria, descripcion, pvp)
                ):
                    print(f'Skipping row with empty fields at position: {column}')
                    column += 1
                    continue

                # Create a new Product object and save it to the database
                product = Product(
                    sku=sku,
                    ean=ean,
                    proveedor=proveedor,
                    categoria=categoria,
                    marca=marca,
                    descripcion=descripcion,
                    pvp=pvp
                )
                product.save()
                print(f'Saved product: {sku}')

                column += 1

            return render(request, 'microservices/upload_success.html')

    else:
        form = ExcelUploadForm()

    print('Invalid form or GET request')
    return render(request, 'microservices/upload.html', {'form': form})